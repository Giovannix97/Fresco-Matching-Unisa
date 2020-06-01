from openpyxl import load_workbook
import xlsxwriter
import os


def createExcel(filename):
    """
        Get the name of the chosen fragment for create file Excel.

        :param image_name:
        :return: A string that represents the excel path.
    """
    folder_path = os.getcwd()
    workbook = xlsxwriter.Workbook(filename)

    os.chdir(folder_path + "\FileExcel")

    workbook.close()
    return folder_path + "\FileExcel" + "\\" + filename

def writeExcel(path,row,column,value,filename):
    """
            :What does he do? open an excel file, write the value in a cell, save the excel file
            :param path_excel_file, value_to_write, filename_file_excel:
            :return: void
    """
    workbook = load_workbook(path)
    sheet = workbook.active

    sheet.cell(row=row, column=column).value = value
    workbook.save(filename)

def readExcel(path,row,column):
    workbook = load_workbook(filename=path)
    sheet = workbook["Sheet1"]
    return sheet.cell(row,column).value



"""
HOW TO USE: SEE THE FOLLOWING
"""
#path = createExcel("Hello.xlsx")
#writeExcel(path,2,"Hello.xlsx")

#value=readExcel(path,2,2)
#print(value)

